import argparse
import datetime
import io
import random
import shutil
from requests import get
from requests.exceptions import RequestException
from contextlib import closing
import bs4
from console import fg, bg, fx, defx
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import requests

GDOCS_URL = {
    'file_id': '1mBB8RAS8uyYG2QoNUAcjMV0VMgFiHyZ7yHBQLRsCmyc',
    'sheet_id': '637442900',
    'format': 'xlsx',
    'template': 'https://docs.google.com/feeds/download/spreadsheets/Export' +
                '?key=%(file_id)s&exportFormat=%(format)s&gid=%(sheet_id)s',
}
DATA = {
    'input_file': None,
    'participants': {},
    'result': {
        'error': None,
        'rank_by_location': {},
        'rank_by_timestamp': {},
    },
}
PROG_NAME = 'GDG Quest Counter'
def main():
    args = parse_args()
    input_file = args.input_file or download_input()
    DATA['input_file'] = input_file

    parse_input(input_file)
    count_quests()

def parse_args():
    parser = argparse.ArgumentParser(prog=PROG_NAME)
    parser.add_argument('-i', '--input-file', dest='input_file',
        help='Input file of participants info (must be an Excel file)')
    return parser.parse_args()

def download_input():
    filepath = 'result.%(format)s' % GDOCS_URL
    url = GDOCS_URL['template'] % GDOCS_URL
    req = requests.get(url)
    with open(filepath, 'w+b') as file:
        file.write(req.content)
    return filepath


def parse_input(input):
    wb = openpyxl.load_workbook(filename=input)
    sh = wb[wb.sheetnames[0]]

    participants = DATA['participants']
    rows_not_processed = []

    row_id = 0
    for row in sh.iter_rows():
        row_id += 1
        if row[0].value == 'Timestamp':
            # Skip header row
            pass
        elif not row[0].is_date:
            rows_not_processed.append(row)
        else:
            person = {
                'row_id': row_id,
                'timestamp': row[0].value,
                'email': row[1].value.strip().lower(),
                'name': row[2].value.strip(),
                'nick_name': row[3].value.strip(),
                'qwiklabs_link': row[4].value.strip(),
                'coursera_link': row[5].value.strip(),
                'location': row[7].value.strip(),
                'quests': [],
                'legal_quests': [],
                'courses': [],
            }
            email = person['email']
            if email in participants:
                # Duplicated entry
                print('Dupplicated input entry ' + person['email'],
                         'at row %d' % row_id)
            participants[email] = person


def count_quests():
    participants = DATA['participants']
    for person in participants.values():
        try:
            count_coursera_of(person)
        except Exception as ex:
            print('Unable to parse QUESTS report for user %s' % person['email'])

    # Track ERROR and OK reports
    error_list = []
    ok_list = []
    for person in participants.values():
        if person.get('error', None):
            error_list.append(person)
        else:
            ok_list.append(person)


def count_coursera_of (person) :
    coursera_link = person['coursera_link']
    if "specialization/certificate" in coursera_link:
        coursera_link = coursera_link.replace('specialization/certificate', 'specialization')
    print('---------------------------')
    print('Processing coursera_link ' + coursera_link + ' ...')

    driver = webdriver.Chrome('C:\\Users\\ADMIN\\Downloads\\chromedriver_win32\\chromedriver.exe')
    driver.get(coursera_link)
    div_all_quests = driver.find_elements_by_class_name('rc-RecordTile')
    course_list = person['courses']
    for div in div_all_quests:
        course_name = div.find_element_by_class_name('course-record-tile-details').find_element_by_tag_name('h3').text.strip()
        course = {
            'course_name': course_name
        }
        course_list.append(course)

    show_quests_report_of(person)
    save_result_txt(person)


def show_quests_report_of(person):
    print('----QUEST REPORT for: \t' + person['name'] + ':')
    print('\tEmail: ' + person['email'])
    print('\tCoursera Link: ' + person['coursera_link'])
    print('\tNumber coursera: ' + str(len(person['courses'])))
    for course in person['courses']:
        print('\t\t - Coursera name: ' + course['course_name'])

def save_result_txt(person):
    with io.open('result_coursera.txt', 'a', encoding='utf-8') as outfile:
        outfile.write('----QUEST REPORT for: \t' + person['name'] + ':\n')
        outfile.write('\tEmail: ' + person['email'] + '\n')
        outfile.write('\tCoursera Link: ' + person['coursera_link'] + '\n')
        outfile.write('\tNumber coursera: ' + str(len(person['courses'])) + '\n')
        for course in person['courses']:
            print('\t\t - Coursera name: ' + course['course_name'] + '\n')
        outfile.write('\n')

if __name__ == '__main__':
    main()












